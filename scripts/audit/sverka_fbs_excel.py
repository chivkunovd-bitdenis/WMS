"""Сверка остатков ФБС — Excel по трём продавцам.

Правила, из-за которых прошлые версии врали:
  * склады только наши («Империя Львов», WMS). Чужие — Фулфилмент 2035877,
    ПакДрайв, ВельветБокс — не наши и в файл не идут;
  * «уехало в заказах» больше нет: это слово склеивало снятия с полки
    с фиктивными возвратами. Теперь три отдельные колонки;
  * возврат на остаток по отменённому заказу — фикция, если поставка уже
    передана WB. Из 298 таких возвратов физическими были 6.
"""
import csv, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STAMP   = "29.08.2026 18:50"
SELLERS = ["ИП Чжоу", "Loviana", "ООО Фэшн"]
S_TAB   = {"ИП Чжоу": "Чжоу", "Loviana": "Loviana", "ООО Фэшн": "Фэшн"}

def rd(p):
    with open(p, encoding="utf-8") as f: return list(csv.DictReader(f))

full = rd("/tmp/out-full.csv")
cab  = [r for r in rd("/tmp/out-cabinet.csv") if r["nash"] == "1"]

wh = collections.defaultdict(dict)
for r in cab: wh[r["seller"]][int(r["wid"])] = r["wh_name"]
WH  = {s: sorted(wh.get(s, {})) for s in SELLERS}
WHN = {s: {w: wh[s][w] for w in WH[s]} for s in SELLERS}
C = collections.defaultdict(lambda: collections.defaultdict(int))
for r in cab: C[(r["seller"], r["sku"])][int(r["wid"])] += int(r["amount"])

wb = Workbook(); wb.remove(wb.active)
HF   = PatternFill("solid", fgColor="1F4E79");  H    = Font(bold=True, color="FFFFFF")
RED  = PatternFill("solid", fgColor="FFC7CE");  GREY = PatternFill("solid", fgColor="D9E1F2")
ORNG = PatternFill("solid", fgColor="FCE4D6")

COLS = ["Артикул", "Привезли на склад", "Сняли с полки",
        "Система вернула фиктивно", "Списали коррекцией", "Отдали другому юрлицу",
        "ОСТАТОК В СИСТЕМЕ", "Фантомов осталось в остатке", "ОСТАТОК БЕЗ ФАНТОМОВ"]

svod = wb.create_sheet("Сводка")
svod.append(["Продавец", "Позиций", "Привезли", "Сняли с полки",
             "Вернули фиктивно", "Списали коррекцией", "Остаток в системе",
             "Фантомов осталось", "ОСТАТОК БЕЗ ФАНТОМОВ", "В кабинете ВБ",
             "Заказов сейчас", "Заказов за всё время", "СВЕРХ ПРИВЕЗЁННОГО",
             "Артикулов с превышением", "Наши склады"])

for s in SELLERS:
    rows_src = [r for r in full if r["seller"] == s]
    tab = wb.create_sheet(S_TAB[s])
    hdr = COLS + [f"В кабинете · {WHN[s][w]}" for w in WH[s]] + \
          ["Всего в кабинете", "Заказов сейчас", "Заказов за всё время", "СВЕРХ ПРИВЕЗЁННОГО"]
    tab.append(hdr)
    for c in tab[1]:
        c.font = H; c.fill = HF
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    rows = []
    for r in rows_src:
        sku = r["sku"]
        privezli, snyali  = int(r["privezli"]), int(r["snyali"])
        fikt,     korr    = int(r["fiktivno"]), int(r["korrekciya"])
        pered,    ostatok = int(r["peredacha"]), int(r["ostatok"])
        zv,       zs      = int(r["zak_vsego"]), int(r["zak_seychas"])
        cb = [C[(s, sku)].get(w, 0) for w in WH[s]]
        if privezli == 0 and ostatok == 0 and zv == 0 and sum(cb) == 0: continue
        fantom = max(0, fikt - korr)          # фикция, ещё не снятая коррекцией
        rows.append([sku, privezli, snyali, fikt, korr, pered,
                     ostatok, fantom, ostatok - fantom] +
                    cb + [sum(cb), zs, zv, max(0, zv - privezli)])

    rows.sort(key=lambda r: (-r[-1], -r[6], r[0]))
    for r in rows:
        tab.append(r)
        if r[-1] > 0:                                        # выставили сверх привезённого
            for j in range(1, len(r) + 1): tab.cell(tab.max_row, j).fill = RED
        elif r[7] > 0:                                       # фантомы в остатке
            for j in range(1, len(r) + 1): tab.cell(tab.max_row, j).fill = ORNG

    tot = ["ИТОГО"] + [sum(r[i] for r in rows) for i in range(1, len(hdr))]
    tab.append([]); tab.append(tot)
    for c in tab[tab.max_row]: c.font = Font(bold=True); c.fill = GREY
    tab.column_dimensions["A"].width = 30
    for i in range(2, len(hdr) + 1): tab.column_dimensions[get_column_letter(i)].width = 15
    tab.freeze_panes = "B2"

    svod.append([s, len(rows), tot[1], tot[2], tot[3], tot[4], tot[6], tot[7], tot[8],
                 tot[9 + len(WH[s])], tot[-3], tot[-2], tot[-1],
                 sum(1 for r in rows if r[-1] > 0),
                 ", ".join(f"{WHN[s][w]} ({w})" for w in WH[s])])

for c in svod[1]:
    c.font = H; c.fill = HF
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
svod.append([])
for line in [
    f"Снято с боевого сервера {STAMP}. Только наши склады — «Империя Львов» и WMS.",
    "",
    "ЧТО ЗНАЧАТ КОЛОНКИ:",
    "Привезли на склад — сколько единиц физически приняли (приёмка + передача прав).",
    "Сняли с полки — сколько раз кладовщик снял товар под заказ. Это факт склада.",
    "Система вернула фиктивно — при отмене заказа система вернула товар на остаток,"
    " хотя поставку уже передали ВБ и товар уехал. Из 298 таких возвратов физическими были 6.",
    "Списали коррекцией — сколько фикции уже снято вручную 29.08.",
    "Фантомов осталось в остатке — фикция минус коррекция. На столько остаток завышен.",
    "ОСТАТОК БЕЗ ФАНТОМОВ — расчётная цифра. Точную даст только пересчёт на складе.",
    "Заказов за всё время — все заказы по нашим складам, включая отменённые."
    " ВБ при отмене единицу в кабинет не возвращает, значит каждый заказ съел выставленный остаток.",
    "СВЕРХ ПРИВЕЗЁННОГО — заказов за всё время минус привезли."
    " Больше нуля — продавец выставил в кабинете больше, чем физически привёз.",
    "",
    "Красным — где выставили сверх привезённого. Оранжевым — где в остатке сидят фантомы.",
]:
    svod.append([line])
for i, w in enumerate([16,9,11,13,15,16,15,15,17,13,13,15,16,15,50], 1):
    svod.column_dimensions[get_column_letter(i)].width = w
svod.freeze_panes = "A2"

out = "/Users/deniscivkunov/Projects/WMS/tmp/sverka-fbs-29-08-1850.xlsx"
wb.save(out); print("готово:", out)
