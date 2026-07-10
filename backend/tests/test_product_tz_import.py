"""Product TZ xlsx import parser and apply."""

# ruff: noqa: RUF001

from __future__ import annotations

import io
import time

import pytest
from httpx import AsyncClient
from openpyxl import Workbook  # type: ignore[import-untyped]

from app.services.product_tz_import_service import (
    ProductTzImportError,
    parse_product_tz_xlsx,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _tz_xlsx_bytes(
    *,
    rows: list[list[object]],
    sheet_name: str = "ТЗ Шаблон (test)",
    merge_tz: tuple[str, str] | None = ("F2", "F4"),
    merge_vendor: tuple[str, str] | None = None,
    tz_text: str = "1. Проверить\n2. Упаковать",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    headers = [
        "Артикул продавца",
        "Фото",
        "Размер",
        "Штрихкод",
        "Информация для этикетки",
        "Пожелания/Инструкция по обработке, упаковке и фасовке",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    if merge_tz is not None and rows:
        ws[merge_tz[0]] = tz_text
        ws.merge_cells(f"{merge_tz[0]}:{merge_tz[1]}")
    if merge_vendor is not None and rows:
        ws.merge_cells(f"{merge_vendor[0]}:{merge_vendor[1]}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_product_tz_xlsx_expands_merged_tz_and_vendor() -> None:
    content = _tz_xlsx_bytes(
        rows=[
            ["ART-beige", None, 46, None, "2031111111111", None],
            [None, None, 48, None, "2031111111112", None],
            [None, None, 50, None, "2031111111113", None],
        ],
        merge_tz=("F2", "F4"),
        merge_vendor=("A2", "A4"),
        tz_text="Общее ТЗ на блок",
    )
    sheet, rows = parse_product_tz_xlsx(content, filename="tz.xlsx")
    assert sheet.startswith("ТЗ Шаблон")
    assert len(rows) == 3
    assert all(r["vendor_article"] == "ART-beige" for r in rows)
    assert all(r["packaging_instructions"] == "Общее ТЗ на блок" for r in rows)
    assert rows[0]["barcode"] == "2031111111111"
    assert rows[0]["size"] == "46"


def test_parse_product_tz_xlsx_sheet_name_is_irrelevant() -> None:
    """Sheet name must not matter — only the column structure does."""
    content = _tz_xlsx_bytes(
        rows=[["ART-any", None, 46, None, "2031111111199", None]],
        sheet_name="Лист1",
        merge_tz=None,
        merge_vendor=None,
        tz_text="",
    )
    sheet, rows = parse_product_tz_xlsx(content, filename="tz.xlsx")
    assert sheet == "Лист1"
    assert len(rows) == 1
    assert rows[0]["vendor_article"] == "ART-any"


def test_parse_product_tz_xlsx_picks_matching_sheet_among_others() -> None:
    """Workbook has an unrelated sheet plus the real template — the real one wins."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Инструкция"
    ws1.append(["Не колонки шаблона", "Просто текст"])
    ws2 = wb.create_sheet("Данные")
    ws2.append(["Артикул продавца", "Фото", "Размер", "Штрихкод", "Информация для этикетки", "ТЗ"])
    ws2.append(["ART-x", None, 46, "2031111111188", None, None])
    buf = io.BytesIO()
    wb.save(buf)
    sheet, rows = parse_product_tz_xlsx(buf.getvalue(), filename="x.xlsx")
    assert sheet == "Данные"
    assert len(rows) == 1
    assert rows[0]["barcode"] == "2031111111188"


def test_parse_product_tz_xlsx_missing_column_in_every_sheet() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Любой лист"
    ws.append(["Что-то другое", "Ещё что-то"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ProductTzImportError) as exc:
        parse_product_tz_xlsx(buf.getvalue(), filename="x.xlsx")
    assert exc.value.code == "missing_column"


@pytest.mark.asyncio
async def test_product_tz_import_preview_and_apply(async_client: AsyncClient) -> None:
    suffix = str(int(time.time() * 1000))
    reg = await async_client.post(
        "/auth/register",
        json={
            "organization_name": "TZ Imp Co",
            "slug": f"tz-imp-{suffix}",
            "admin_email": f"tz-imp-{suffix}@example.com",
            "password": "password123",
        },
    )
    assert reg.status_code == 200, reg.text
    token = str(reg.json()["access_token"])
    h = {"Authorization": f"Bearer {token}"}

    seller = await async_client.post("/sellers", headers=h, json={"name": "Seller TZ"})
    assert seller.status_code in {200, 201}, seller.text
    seller_id = seller.json()["id"]

    content = _tz_xlsx_bytes(
        rows=[
            ["Chin-56005beige", None, 46, None, "2038493603840", None],
            ["Chin-56005beige", None, 48, None, "2038493603857", None],
        ],
        merge_tz=("F2", "F3"),
        tz_text="1 - Достаем куртку",
    )

    preview = await async_client.post(
        "/products/import-tz/preview",
        headers=h,
        data={"seller_id": seller_id},
        files={"file": ("tz.xlsx", content, XLSX_MIME)},
    )
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["summary"]["create_count"] == 2
    assert body["summary"]["error_count"] == 0
    assert all(r["packaging_instructions"] == "1 - Достаем куртку" for r in body["rows"])

    apply = await async_client.post(
        "/products/import-tz/apply",
        headers=h,
        data={"seller_id": seller_id, "ignore_errors": "false"},
        files={"file": ("tz.xlsx", content, XLSX_MIME)},
    )
    assert apply.status_code == 200, apply.text
    applied = apply.json()
    assert applied["created_count"] == 2

    catalog = await async_client.get(
        f"/products/ff-catalog?seller_id={seller_id}",
        headers=h,
    )
    assert catalog.status_code == 200
    rows = catalog.json()
    assert len(rows) == 2
    assert all(r["is_manual"] is True for r in rows)
    assert all(r["has_packaging_instructions"] is True for r in rows)
    barcodes = {r["wb_primary_barcode"] for r in rows}
    assert barcodes == {"2038493603840", "2038493603857"}

    content2 = _tz_xlsx_bytes(
        rows=[
            ["Chin-56005beige", None, 46, None, "2038493603840", None],
            ["Chin-56005beige", None, 48, None, "2038493603857", None],
        ],
        merge_tz=("F2", "F3"),
        tz_text="Обновлённое ТЗ",
    )
    apply2 = await async_client.post(
        "/products/import-tz/apply",
        headers=h,
        data={"seller_id": seller_id, "ignore_errors": "false"},
        files={"file": ("tz.xlsx", content2, XLSX_MIME)},
    )
    assert apply2.status_code == 200, apply2.text
    assert apply2.json()["updated_count"] == 2
    assert apply2.json()["created_count"] == 0

    catalog2 = await async_client.get(
        f"/products/ff-catalog?seller_id={seller_id}",
        headers=h,
    )
    assert all(
        "Обновлённое ТЗ" in (r["packaging_instructions"] or "") for r in catalog2.json()
    )


@pytest.mark.asyncio
async def test_create_product_null_dims_and_fields(async_client: AsyncClient) -> None:
    suffix = str(int(time.time() * 1000))
    reg = await async_client.post(
        "/auth/register",
        json={
            "organization_name": "Manual Co",
            "slug": f"manual-{suffix}",
            "admin_email": f"manual-{suffix}@example.com",
            "password": "password123",
        },
    )
    assert reg.status_code == 200
    token = str(reg.json()["access_token"])
    h = {"Authorization": f"Bearer {token}"}
    seller = await async_client.post("/sellers", headers=h, json={"name": "S"})
    assert seller.status_code in {200, 201}, seller.text
    seller_id = seller.json()["id"]

    pr = await async_client.post(
        "/products",
        headers=h,
        json={
            "name": "Ручной товар",
            "sku_code": f"MAN-{suffix}",
            "seller_id": seller_id,
            "wb_barcode": f"204{suffix[-10:]}",
            "wb_size": "46",
            "wb_vendor_code": "ART-manual",
            "packaging_instructions": "Сложить и проклеить",
        },
    )
    assert pr.status_code == 200, pr.text
    data = pr.json()
    assert data["length_mm"] is None
    assert data["width_mm"] is None
    assert data["height_mm"] is None
    assert data["volume_liters"] is None
    assert data["is_manual"] is True
    assert data["wb_barcode"].startswith("204")
    assert data["wb_size"] == "46"


@pytest.mark.asyncio
async def test_tz_import_cross_seller_barcode_in_preview(
    async_client: AsyncClient,
) -> None:
    suffix = str(int(time.time() * 1000))
    reg = await async_client.post(
        "/auth/register",
        json={
            "organization_name": "Cross BC",
            "slug": f"cross-bc-{suffix}",
            "admin_email": f"cross-bc-{suffix}@example.com",
            "password": "password123",
        },
    )
    assert reg.status_code == 200
    token = str(reg.json()["access_token"])
    h = {"Authorization": f"Bearer {token}"}
    sa = await async_client.post("/sellers", headers=h, json={"name": "A"})
    sb = await async_client.post("/sellers", headers=h, json={"name": "B"})
    seller_a = sa.json()["id"]
    seller_b = sb.json()["id"]
    barcode = f"2038{suffix[-9:]}"

    created = await async_client.post(
        "/products",
        headers=h,
        json={
            "name": "Owned by A",
            "sku_code": f"A-{suffix}",
            "seller_id": seller_a,
            "wb_barcode": barcode,
        },
    )
    assert created.status_code == 200, created.text

    content = _tz_xlsx_bytes(
        rows=[["ART-B", None, 46, None, barcode, None]],
        merge_tz=None,
        tz_text="",
    )
    # put TZ in cell without merge
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "ТЗ Шаблон (cross)"
    ws.append(
        [
            "Артикул продавца",
            "Фото",
            "Размер",
            "Штрихкод",
            "Информация для этикетки",
            "Пожелания/Инструкция по обработке, упаковке и фасовке",
        ]
    )
    ws.append(["ART-B", None, 46, None, barcode, "TZ"])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    preview = await async_client.post(
        "/products/import-tz/preview",
        headers=h,
        data={"seller_id": seller_b},
        files={"file": ("tz.xlsx", content, XLSX_MIME)},
    )
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["summary"]["error_count"] == 1
    assert body["rows"][0]["error_code"] == "barcode_taken_other_seller"
    assert body["summary"]["create_count"] == 0
