from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.db.session import SessionLocal
from app.models.inbound_intake import InboundIntakeLine, InboundIntakeRequest
from app.models.inventory_balance import InventoryBalance
from app.models.inventory_movement import InventoryMovement
from app.models.inventory_reservation import InventoryReservation
from app.models.marking_code import MarkingCode, MarkingCodeEvent
from app.models.outbound_shipment import OutboundShipmentLine, OutboundShipmentRequest
from app.models.product import Product
from app.models.seller import Seller
from app.models.storage_location import StorageLocation
from app.models.tenant import Tenant
from app.models.warehouse import Warehouse
from app.services import inventory_service
from app.services.ownership_transfer_service import (
    MOVEMENT_TYPE_OWNERSHIP_IN,
    MOVEMENT_TYPE_OWNERSHIP_OUT,
    MOVEMENT_TYPE_OWNERSHIP_RECEIPT,
    OwnershipTransferError,
    OwnershipTransferInput,
    apply_ownership_transfer_plan,
    build_ownership_transfer_plan,
)
from app.services.sorting_location_service import SORTING_LOCATION_CODE
from scripts.transfer_loviana_fashion_20260825 import parse_transfer_workbook


async def _seed_transfer_case() -> dict[str, uuid.UUID]:
    async with SessionLocal() as session:
        tenant = Tenant(name="Transfer tenant", slug=f"transfer-{uuid.uuid4().hex}")
        session.add(tenant)
        await session.flush()
        warehouse = Warehouse(
            tenant_id=tenant.id,
            name="Основной склад",
            code="MAIN",
            barcode=f"WH-{uuid.uuid4().hex[:12]}",
        )
        source_seller = Seller(tenant_id=tenant.id, name="Loviana")
        target_seller = Seller(tenant_id=tenant.id, name="ООО Фэшн")
        session.add_all([warehouse, source_seller, target_seller])
        await session.flush()
        location = StorageLocation(
            tenant_id=tenant.id,
            warehouse_id=warehouse.id,
            code="A-01-01",
            barcode=f"LOC-{uuid.uuid4().hex[:12]}",
        )
        source_product = Product(
            tenant_id=tenant.id,
            seller_id=source_seller.id,
            name="Товар 1 Loviana",
            sku_code="P1/36",
            wb_barcode="111",
        )
        target_product = Product(
            tenant_id=tenant.id,
            seller_id=target_seller.id,
            name="Товар 1 Фэшн",
            sku_code="P1/36",
            wb_barcode="211",
        )
        receipt_product = Product(
            tenant_id=tenant.id,
            seller_id=target_seller.id,
            name="Товар 2 Фэшн",
            sku_code="P2/36",
            wb_barcode="212",
        )
        session.add_all([location, source_product, target_product, receipt_product])
        await session.flush()
        await inventory_service.record_movement_and_adjust_balance(
            session,
            tenant_id=tenant.id,
            product_id=source_product.id,
            storage_location_id=location.id,
            quantity_delta=5,
            movement_type="test_seed",
        )
        outbound = OutboundShipmentRequest(
            tenant_id=tenant.id,
            warehouse_id=warehouse.id,
            seller_id=source_seller.id,
            status="draft",
        )
        session.add(outbound)
        await session.flush()
        outbound_line = OutboundShipmentLine(
            request_id=outbound.id,
            product_id=source_product.id,
            quantity=2,
            shipped_qty=0,
            storage_location_id=location.id,
        )
        session.add(outbound_line)
        await session.flush()
        session.add(
            InventoryReservation(
                tenant_id=tenant.id,
                outbound_shipment_line_id=outbound_line.id,
                product_id=source_product.id,
                storage_location_id=location.id,
                warehouse_id=None,
                quantity=2,
            )
        )
        marking = MarkingCode(
            tenant_id=tenant.id,
            seller_id=source_seller.id,
            product_id=source_product.id,
            cis_code=f"010460000000000021{uuid.uuid4().hex}",
            status="available",
        )
        reserved_marking = MarkingCode(
            tenant_id=tenant.id,
            seller_id=source_seller.id,
            product_id=source_product.id,
            cis_code=f"010460000000000021{uuid.uuid4().hex}",
            status="reserved",
        )
        session.add_all([marking, reserved_marking])
        await session.commit()
        return {
            "tenant": tenant.id,
            "warehouse": warehouse.id,
            "location": location.id,
            "source_seller": source_seller.id,
            "target_seller": target_seller.id,
            "source_product": source_product.id,
            "target_product": target_product.id,
            "receipt_product": receipt_product.id,
            "marking": marking.id,
            "reserved_marking": reserved_marking.id,
        }


def _inputs() -> list[OwnershipTransferInput]:
    return [
        OwnershipTransferInput(
            row_number=30,
            sku="P1",
            size="36",
            source_barcode="111",
            target_barcode="211",
            quantity=7,
        ),
        OwnershipTransferInput(
            row_number=40,
            sku="P2",
            size="36",
            source_barcode=None,
            target_barcode="212",
            quantity=4,
        ),
    ]


@pytest.mark.asyncio
async def test_transfer_is_atomic_keeps_cell_receives_shortage_and_is_idempotent(
    async_client: Any,
) -> None:
    ids = await _seed_transfer_case()
    async with SessionLocal() as session:
        plan = await build_ownership_transfer_plan(
            session,
            tenant_id=ids["tenant"],
            warehouse_id=ids["warehouse"],
            run_id="transfer-24.08-test-1",
            source_sha256="a" * 64,
            inputs=_inputs(),
        )
        assert plan.blockers == []
        assert plan.transfer_quantity == 3
        assert plan.receipt_quantity == 8
        assert plan.rows[0].source_before == 5
        assert plan.rows[0].source_reserved == 2
        assert plan.rows[0].allocations[0].location_id == ids["location"]
        assert plan.rows[0].allocations[0].quantity == 3
        assert plan.rows[0].receipt_quantity == 4
        assert len(plan.rows[0].marking_code_ids) == 1
        assert plan.rows[0].marking_codes_blocked == 1

        request = await apply_ownership_transfer_plan(
            session,
            approved_token=plan.approval_token,
            plan=plan,
        )
        await session.commit()
        request_id = request.id

    async with SessionLocal() as session:
        source_balance = await session.scalar(
            select(InventoryBalance).where(
                InventoryBalance.product_id == ids["source_product"],
                InventoryBalance.storage_location_id == ids["location"],
            )
        )
        target_balance = await session.scalar(
            select(InventoryBalance).where(
                InventoryBalance.product_id == ids["target_product"],
                InventoryBalance.storage_location_id == ids["location"],
            )
        )
        sorting = await session.scalar(
            select(StorageLocation).where(
                StorageLocation.warehouse_id == ids["warehouse"],
                StorageLocation.code == SORTING_LOCATION_CODE,
            )
        )
        assert source_balance is not None and source_balance.quantity == 2
        assert target_balance is not None and target_balance.quantity == 3
        assert sorting is not None
        receipt_balances = {
            balance.product_id: balance.quantity
            for balance in (
                await session.execute(
                    select(InventoryBalance).where(
                        InventoryBalance.storage_location_id == sorting.id
                    )
                )
            ).scalars()
        }
        assert receipt_balances[ids["target_product"]] == 4
        assert receipt_balances[ids["receipt_product"]] == 4

        movements = list(
            (
                await session.execute(
                    select(InventoryMovement).where(
                        InventoryMovement.movement_type.in_(
                            {
                                MOVEMENT_TYPE_OWNERSHIP_OUT,
                                MOVEMENT_TYPE_OWNERSHIP_IN,
                                MOVEMENT_TYPE_OWNERSHIP_RECEIPT,
                            }
                        )
                    )
                )
            ).scalars()
        )
        pair = [movement for movement in movements if movement.transfer_group_id is not None]
        assert len(pair) == 2
        assert pair[0].transfer_group_id == pair[1].transfer_group_id
        assert sum(m.quantity_delta for m in movements) == 8

        inbound = await session.get(InboundIntakeRequest, request_id)
        assert inbound is not None
        assert inbound.status == "sorting"
        assert "Transfer/24.08" in (inbound.waybill_number or "")
        inbound_lines = list(
            (
                await session.execute(
                    select(InboundIntakeLine).where(InboundIntakeLine.request_id == request_id)
                )
            ).scalars()
        )
        assert sorted(line.actual_qty for line in inbound_lines) == [4, 4]

        marking = await session.get(MarkingCode, ids["marking"])
        assert marking is not None
        assert marking.seller_id == ids["target_seller"]
        assert marking.product_id == ids["target_product"]
        event = await session.scalar(
            select(MarkingCodeEvent).where(MarkingCodeEvent.code_id == marking.id)
        )
        assert event is not None and event.event_type == "transferred"
        reserved_marking = await session.get(MarkingCode, ids["reserved_marking"])
        assert reserved_marking is not None
        assert reserved_marking.seller_id == ids["source_seller"]
        assert reserved_marking.product_id == ids["source_product"]

        repeated = await build_ownership_transfer_plan(
            session,
            tenant_id=ids["tenant"],
            warehouse_id=ids["warehouse"],
            run_id="transfer-24.08-test-1",
            source_sha256="a" * 64,
            inputs=_inputs(),
        )
        assert "run_already_applied" in repeated.blockers


@pytest.mark.asyncio
async def test_stale_dry_run_token_does_not_write_anything(async_client: Any) -> None:
    ids = await _seed_transfer_case()
    async with SessionLocal() as session:
        dry_plan = await build_ownership_transfer_plan(
            session,
            tenant_id=ids["tenant"],
            warehouse_id=ids["warehouse"],
            run_id="transfer-24.08-test-stale",
            source_sha256="b" * 64,
            inputs=_inputs(),
        )
        old_token = dry_plan.approval_token
        await inventory_service.record_movement_and_adjust_balance(
            session,
            tenant_id=ids["tenant"],
            product_id=ids["target_product"],
            storage_location_id=ids["location"],
            quantity_delta=1,
            movement_type="state_changed_after_dry_run",
        )
        await session.commit()

    async with SessionLocal() as session:
        current_plan = await build_ownership_transfer_plan(
            session,
            tenant_id=ids["tenant"],
            warehouse_id=ids["warehouse"],
            run_id="transfer-24.08-test-stale",
            source_sha256="b" * 64,
            inputs=_inputs(),
            lock=True,
        )
        assert current_plan.approval_token != old_token
        with pytest.raises(OwnershipTransferError, match="approval_token_mismatch"):
            await apply_ownership_transfer_plan(
                session,
                approved_token=old_token,
                plan=current_plan,
            )
        await session.rollback()

    async with SessionLocal() as session:
        marker = await session.get(InboundIntakeRequest, current_plan.marker_request_id)
        assert marker is None


@pytest.mark.asyncio
async def test_missing_target_product_blocks_whole_plan(async_client: Any) -> None:
    ids = await _seed_transfer_case()
    bad_inputs = [
        OwnershipTransferInput(
            row_number=30,
            sku="P1",
            size="36",
            source_barcode="111",
            target_barcode="missing",
            quantity=1,
        )
    ]
    async with SessionLocal() as session:
        plan = await build_ownership_transfer_plan(
            session,
            tenant_id=ids["tenant"],
            warehouse_id=ids["warehouse"],
            run_id="transfer-24.08-test-missing",
            source_sha256="c" * 64,
            inputs=bad_inputs,
        )
        assert "row_30:target_product_not_found" in plan.blockers
        with pytest.raises(OwnershipTransferError, match="plan_has_blockers"):
            await apply_ownership_transfer_plan(
                session,
                approved_token=plan.approval_token,
                plan=plan,
            )


def test_workbook_parser_excludes_all_j308_and_corrects_row_69(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "24.08"
    sheet.append(
        [
            None,
            None,
            "Артикул продавца",
            "Размер",
            None,
            None,
            "Баркод Loviana",
            "Баркод Fashion",
            "Переносим шт.",
        ]
    )
    sheet.cell(25, 3, "P1")
    sheet.cell(25, 4, 36)
    sheet.cell(25, 7, 111)
    sheet.cell(25, 8, 211)
    sheet.cell(25, 9, 10)
    sheet.cell(69, 3, "F907-1")
    sheet.cell(69, 4, 40)
    sheet.cell(69, 7, 2054452452250)
    sheet.cell(69, 8, 2041885503642)
    sheet.cell(69, 9, 110)
    for row_number, sku in ((84, "J308-6"), (90, "J308-25"), (96, "J308-24")):
        sheet.cell(row_number, 3, sku)
        sheet.cell(row_number, 4, 36)
        sheet.cell(row_number, 7, row_number * 10)
        sheet.cell(row_number, 8, row_number * 10 + 1)
        sheet.cell(row_number, 9, 20)
    path = tmp_path / "Transfer.xlsx"
    workbook.save(path)

    _checksum, inputs = parse_transfer_workbook(
        path,
        enforce_known_source=False,
        enforce_expected_scope=False,
    )
    j308 = [item for item in inputs if item.sku.startswith("J308")]
    assert len(j308) == 3
    assert {item.excluded_reason for item in j308} == {
        "j308_loafers_excluded_by_owner"
    }
    corrected = next(item for item in inputs if item.row_number == 69)
    assert corrected.original_target_barcode == "2041885503642"
    assert corrected.target_barcode == "2041373819071"
