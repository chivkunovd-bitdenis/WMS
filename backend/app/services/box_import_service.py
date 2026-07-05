"""Parse «Штрих-код комбайн» xlsx exports into inbound/MP box layouts."""

from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.inbound_intake import InboundIntakeBox
from app.models.marketplace_unload import MarketplaceUnloadBox
from app.services import inbound_intake_box_service as inbound_box_svc
from app.services import marketplace_unload_box_service as mp_box_svc
from app.services.seller_wb_catalog_service import list_seller_wb_catalog_rows

DEFAULT_MP_BOX_PRESET = "60_40_40"

REQUIRED_COLUMNS: dict[str, str] = {
    "штрих-код": "barcode",
    "кол-во": "quantity",
    "адрес": "address",
}


class BoxImportError(Exception):
    def __init__(self, code: str, message: str = "") -> None:
        super().__init__(message or code)
        self.code = code
        self.message = message or code


@dataclass(frozen=True)
class ProductCatalogRow:
    product_id: uuid.UUID
    sku_code: str
    product_name: str
    wb_primary_barcode: str | None
    wb_barcodes: tuple[str, ...]


@dataclass(frozen=True)
class BoxImportLinePreview:
    barcode: str
    product_id: uuid.UUID | None
    sku_code: str | None
    product_name: str | None
    quantity: int


@dataclass(frozen=True)
class BoxImportBoxPreview:
    address: str
    lines: tuple[BoxImportLinePreview, ...]
    total_qty: int


@dataclass(frozen=True)
class BoxImportRowError:
    row: int
    barcode: str | None
    code: str
    message: str


@dataclass(frozen=True)
class BoxImportPreviewSummary:
    boxes_count: int
    positions: int
    total_units: int
    error_count: int


@dataclass(frozen=True)
class BoxImportPreviewResult:
    boxes: tuple[BoxImportBoxPreview, ...]
    errors: tuple[BoxImportRowError, ...]
    summary: BoxImportPreviewSummary


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"\s+", " ", text)


def _parse_quantity(raw: Any, row_num: int) -> int:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raise BoxImportError("invalid_quantity", f"row {row_num}: empty quantity")
    if isinstance(raw, bool):
        raise BoxImportError("invalid_quantity", f"row {row_num}: invalid quantity")
    if isinstance(raw, int):
        qty = raw
    elif isinstance(raw, float):
        if not raw.is_integer():
            raise BoxImportError("invalid_quantity", f"row {row_num}: fractional quantity")
        qty = int(raw)
    else:
        text = str(raw).strip().replace(",", ".")
        try:
            num = float(text)
        except ValueError as exc:
            raise BoxImportError(
                "invalid_quantity", f"row {row_num}: invalid quantity"
            ) from exc
        if not num.is_integer():
            raise BoxImportError("invalid_quantity", f"row {row_num}: fractional quantity")
        qty = int(num)
    if qty < 1:
        raise BoxImportError("invalid_quantity", f"row {row_num}: quantity must be >= 1")
    return qty


def _resolve_product_id(catalog: list[ProductCatalogRow], barcode: str) -> ProductCatalogRow | None:
    code = barcode.strip()
    if not code:
        return None
    lower = code.lower()
    for row in catalog:
        if row.sku_code.lower() == lower:
            return row
        primary = (row.wb_primary_barcode or "").strip()
        if primary and primary.lower() == lower:
            return row
        for wb in row.wb_barcodes:
            if wb.strip().lower() == lower:
                return row
    return None


def parse_box_import_xlsx(
    content: bytes,
    *,
    filename: str,
    catalog: list[ProductCatalogRow],
) -> BoxImportPreviewResult:
    lower_name = filename.lower()
    if not lower_name.endswith(".xlsx"):
        raise BoxImportError("unsupported_file_type")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BoxImportError("unsupported_file_type") from exc

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise BoxImportError("empty_file") from exc

    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        header_key = _normalize_header(cell)
        if header_key in REQUIRED_COLUMNS:
            col_map[REQUIRED_COLUMNS[header_key]] = idx

    for required_key, field in REQUIRED_COLUMNS.items():
        if field not in col_map:
            raise BoxImportError("missing_column", required_key)

    aggregated: dict[tuple[str, str], int] = {}
    errors: list[BoxImportRowError] = []
    row_num = 1

    for row in rows_iter:
        row_num += 1
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_values = row
        barcode_idx = col_map["barcode"]
        qty_idx = col_map["quantity"]
        address_idx = col_map["address"]
        barcode_raw = row_values[barcode_idx] if barcode_idx < len(row_values) else None
        address_raw = row_values[address_idx] if address_idx < len(row_values) else None
        qty_raw = row_values[qty_idx] if qty_idx < len(row_values) else None
        barcode = "" if barcode_raw is None else str(barcode_raw).strip()
        address = "" if address_raw is None else str(address_raw).strip()

        if not barcode:
            errors.append(
                BoxImportRowError(
                    row=row_num,
                    barcode=None,
                    code="empty_barcode",
                    message=f"строка {row_num}: пустой штрих-код",
                )
            )
            continue
        if not address:
            errors.append(
                BoxImportRowError(
                    row=row_num,
                    barcode=barcode,
                    code="empty_address",
                    message=f"строка {row_num}: пустой адрес",
                )
            )
            continue
        try:
            qty = _parse_quantity(qty_raw, row_num)
        except BoxImportError as exc:
            errors.append(
                BoxImportRowError(
                    row=row_num,
                    barcode=barcode,
                    code=exc.code,
                    message=exc.message,
                )
            )
            continue

        key = (barcode, address)
        aggregated[key] = aggregated.get(key, 0) + qty

    boxes_map: dict[str, list[BoxImportLinePreview]] = {}
    for (barcode, address), quantity in aggregated.items():
        product = _resolve_product_id(catalog, barcode)
        if product is None:
            errors.append(
                BoxImportRowError(
                    row=0,
                    barcode=barcode,
                    code="barcode_not_found",
                    message=f"ШК {barcode} не найден",
                )
            )
            line = BoxImportLinePreview(
                barcode=barcode,
                product_id=None,
                sku_code=None,
                product_name=None,
                quantity=quantity,
            )
        else:
            line = BoxImportLinePreview(
                barcode=barcode,
                product_id=product.product_id,
                sku_code=product.sku_code,
                product_name=product.product_name,
                quantity=quantity,
            )
        boxes_map.setdefault(address, []).append(line)

    boxes: list[BoxImportBoxPreview] = []
    total_units = 0
    positions = 0
    for address in sorted(boxes_map.keys(), key=lambda a: (len(a), a)):
        lines = boxes_map[address]
        box_total = sum(ln.quantity for ln in lines)
        total_units += box_total
        positions += len(lines)
        boxes.append(
            BoxImportBoxPreview(
                address=address,
                lines=tuple(lines),
                total_qty=box_total,
            )
        )

    summary = BoxImportPreviewSummary(
        boxes_count=len(boxes),
        positions=positions,
        total_units=total_units,
        error_count=len(errors),
    )
    return BoxImportPreviewResult(
        boxes=tuple(boxes),
        errors=tuple(errors),
        summary=summary,
    )


async def build_product_catalog(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    seller_id: uuid.UUID,
) -> list[ProductCatalogRow]:
    rows = await list_seller_wb_catalog_rows(session, tenant_id, seller_id)
    return [
        ProductCatalogRow(
            product_id=row.product_id,
            sku_code=row.sku_code,
            product_name=row.name,
            wb_primary_barcode=row.wb_primary_barcode,
            wb_barcodes=row.wb_barcodes,
        )
        for row in rows
    ]


async def apply_inbound_box_import(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    request_id: uuid.UUID,
    preview: BoxImportPreviewResult,
    *,
    ignore_errors: bool = False,
) -> tuple[list[InboundIntakeBox], list[BoxImportRowError]]:
    if preview.errors and not ignore_errors:
        raise BoxImportError("row_errors")

    applied_errors = list(preview.errors)
    created_boxes: list[InboundIntakeBox] = []
    address_to_box: dict[str, uuid.UUID] = {}

    for box_preview in preview.boxes:
        resolvable_lines = [ln for ln in box_preview.lines if ln.product_id is not None]
        if not resolvable_lines:
            continue
        box = await inbound_box_svc.create_open_box(session, tenant_id, request_id)
        created_boxes.append(box)
        address_to_box[box_preview.address] = box.id
        for line in resolvable_lines:
            assert line.product_id is not None
            await inbound_box_svc.set_product_quantity_in_open_box(
                session,
                tenant_id,
                request_id,
                box.id,
                product_id=line.product_id,
                quantity=line.quantity,
            )

    await session.commit()
    return created_boxes, applied_errors


async def apply_marketplace_box_import(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    request_id: uuid.UUID,
    preview: BoxImportPreviewResult,
    *,
    ignore_errors: bool = False,
    box_preset: str = DEFAULT_MP_BOX_PRESET,
) -> tuple[list[MarketplaceUnloadBox], list[BoxImportRowError]]:
    if preview.errors and not ignore_errors:
        raise BoxImportError("row_errors")

    applied_errors = list(preview.errors)
    boxes_to_fill: list[tuple[BoxImportBoxPreview, list[BoxImportLinePreview]]] = []
    for box_preview in preview.boxes:
        resolvable_lines = [ln for ln in box_preview.lines if ln.product_id is not None]
        if resolvable_lines:
            boxes_to_fill.append((box_preview, resolvable_lines))

    if not boxes_to_fill:
        return [], applied_errors

    created_boxes = await mp_box_svc.create_boxes_batch(
        session,
        tenant_id,
        request_id,
        count=len(boxes_to_fill),
        box_preset=box_preset,
    )
    for (_preview_box, lines), box in zip(boxes_to_fill, created_boxes, strict=True):
        for line in lines:
            assert line.product_id is not None
            await mp_box_svc.add_manual_qty_to_box(
                session,
                tenant_id,
                box.id,
                product_id=line.product_id,
                storage_location_id=None,
                quantity=line.quantity,
            )

    return created_boxes, applied_errors
