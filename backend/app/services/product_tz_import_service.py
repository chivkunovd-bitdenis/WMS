# ruff: noqa: RUF001
"""Parse FF TZ template xlsx and create/update catalog products."""

from __future__ import annotations

import hashlib
import io
import re
import uuid
from dataclasses import dataclass
from typing import Any, Literal

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.product_tz_import import ProductTzImport
from app.models.seller import Seller
from app.services.catalog_service import (
    CatalogError,
    create_product,
    update_packaging_instructions,
)
from app.services.wb_card_enrichment import WbSizeVariant, sku_code_for_wb_variant

_BARCODE_RE = re.compile(r"^\d{8,32}$")
_IMPORT_TYPE = "product_tz"
_NO_WAREHOUSE_SCOPE = "none"
_IDEMPOTENCY_CONSTRAINT = "uq_product_tz_import_scope_hash"

HEADER_ALIASES: dict[str, str] = {
    "название": "name",
    "название товара": "name",
    "товар": "name",
    "артикул продавца": "vendor_article",
    "sku": "sku",
    "внутренний sku": "sku",
    "артикул sku": "sku",
    "артикул wb": "wb_nm_id",
    "артикул wildberries": "wb_nm_id",
    "wb/nmid": "wb_nm_id",
    "wb nm id": "wb_nm_id",
    "nmid": "wb_nm_id",
    "размер": "size",
    "штрихкод": "barcode",
    "шк": "barcode",
    "информация для этикетки": "label_barcode",
    "тз упаковки": "tz",
    "пожелания/инструкция по обработке, упаковке и фасовке": "tz",
    "пожелания/инструкция по обработке упаковке и фасовке": "tz",
}

_EXPAND_FIELDS = (
    "name",
    "vendor_article",
    "sku",
    "wb_nm_id",
    "size",
    "barcode",
    "label_barcode",
    "tz",
)


class ProductTzImportError(Exception):
    def __init__(self, code: str, message: str = "") -> None:
        super().__init__(message or code)
        self.code = code
        self.message = message or code


def _is_idempotency_conflict(exc: IntegrityError) -> bool:
    diag = getattr(exc.orig, "diag", None)
    if getattr(diag, "constraint_name", None) == _IDEMPOTENCY_CONSTRAINT:
        return True
    message = str(exc.orig).lower()
    if _IDEMPOTENCY_CONSTRAINT in message:
        return True
    sqlite_columns = (
        "product_tz_imports.tenant_id, "
        "product_tz_imports.seller_id, "
        "product_tz_imports.warehouse_scope, "
        "product_tz_imports.import_type, "
        "product_tz_imports.file_sha256"
    )
    return "unique constraint failed:" in message and sqlite_columns in message


@dataclass(frozen=True)
class ProductTzRowError:
    row: int
    barcode: str | None
    code: str
    message: str


@dataclass(frozen=True)
class ProductTzRowPreview:
    row: int
    wb_nm_id: int | None
    vendor_article: str | None
    size: str | None
    barcode: str | None
    name: str
    sku_code: str
    packaging_instructions: str | None
    declared_quantity: int | None
    action: Literal["create", "update", "skip", "error"]
    product_id: uuid.UUID | None
    error_code: str | None
    error_message: str | None


@dataclass(frozen=True)
class ProductTzPreviewSummary:
    total: int
    create_count: int
    update_count: int
    skip_count: int
    error_count: int
    declared_total: int


@dataclass(frozen=True)
class ProductTzPreviewResult:
    rows: tuple[ProductTzRowPreview, ...]
    errors: tuple[ProductTzRowError, ...]
    summary: ProductTzPreviewSummary
    sheet_name: str


@dataclass(frozen=True)
class ProductTzApplyResult:
    created_count: int
    updated_count: int
    skipped_count: int
    product_ids: tuple[uuid.UUID, ...]
    summary: ProductTzPreviewSummary
    errors: tuple[ProductTzRowError, ...]
    added_quantity: int
    movement_count: int
    already_applied: bool
    warehouse_id: uuid.UUID | None


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("\n", " ")
    return re.sub(r"\s+", " ", text)


def _cell_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _as_barcode(value: object) -> str | None:
    text = _cell_str(value)
    if text is None:
        return None
    compact = re.sub(r"\s+", "", text)
    if _BARCODE_RE.match(compact):
        return compact
    return None


def _parse_wb_nm_id(value: object) -> tuple[int | None, str | None]:
    text = _cell_str(value)
    if text is None:
        return None, None
    compact = re.sub(r"\s+", "", text)
    if compact.isdigit():
        return int(compact), None
    return None, "WB/nmId должен быть числом."


def build_product_tz_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог товаров"
    ws.append(
        [
            "Название товара",
            "Артикул продавца",
            "SKU",
            "Штрихкод",
            "WB/nmId",
            "Размер",
            "ТЗ упаковки",
        ]
    )
    ws.append(
        [
            "Футболка oversize",
            "ART-001",
            "ART-001/46",
            "2040000000001",
            "123456789",
            "46",
            "Проверить пакет и наклеить товарный ШК",
        ]
    )
    widths = {
        "A": 28,
        "B": 22,
        "C": 22,
        "D": 18,
        "E": 16,
        "F": 12,
        "G": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _find_header_row(ws: Worksheet) -> tuple[int, dict[str, int]]:
    max_scan = min(ws.max_row or 1, 30)
    for r in range(1, max_scan + 1):
        mapping: dict[str, int] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            key = HEADER_ALIASES.get(_norm_header(ws.cell(r, c).value))
            if key and key not in mapping:
                mapping[key] = c
        if "vendor_article" in mapping and ("barcode" in mapping or "label_barcode" in mapping):
            return r, mapping
    raise ProductTzImportError(
        "missing_column",
        "В файле нет обязательных колонок: «Артикул продавца» и штрихкод/этикетка.",
    )


def _find_tz_sheet(wb: Any) -> tuple[str, int, dict[str, int]]:
    """Pick the sheet that matches the TZ template structure.

    Sheet name is irrelevant — every sheet in the workbook is scanned in order
    and the first one whose header row contains the required columns wins.
    """
    last_error: ProductTzImportError | None = None
    for name in wb.sheetnames:
        try:
            header_row, cols = _find_header_row(wb[name])
        except ProductTzImportError as exc:
            last_error = exc
            continue
        return name, header_row, cols
    raise last_error or ProductTzImportError(
        "missing_column",
        "В файле нет обязательных колонок: «Артикул продавца» и штрихкод/этикетка.",
    )


def _merged_values_by_row(
    ws: Worksheet, col: int, header_row: int
) -> dict[int, str]:
    """Expand Excel merged cells for one column onto every covered data row."""
    out: dict[int, str] = {}
    covered: set[int] = set()
    for mr in ws.merged_cells.ranges:
        if not (mr.min_col <= col <= mr.max_col):
            continue
        text = _cell_str(ws.cell(mr.min_row, col).value)
        if not text:
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            if r <= header_row:
                continue
            out[r] = text
            covered.add(r)
    max_row = ws.max_row or header_row
    for r in range(header_row + 1, max_row + 1):
        if r in covered:
            continue
        text = _cell_str(ws.cell(r, col).value)
        if text:
            out[r] = text
    return out


def _resolve_barcode(*, barcode_raw: str | None, label_raw: str | None) -> str | None:
    """Prefer «Штрихкод», then digit-looking «Информация для этикетки»."""
    for candidate in (barcode_raw, label_raw):
        parsed = _as_barcode(candidate)
        if parsed:
            return parsed
    return None


def _sku_for_row(*, vendor: str, size: str | None, barcode: str) -> str:
    variant = WbSizeVariant(chrt_id=None, size_label=size, barcode=barcode)
    return sku_code_for_wb_variant(vendor, None, variant, multi_variant=True)


def _display_name(name: str | None, vendor: str | None) -> str:
    clean = (name or "").strip()
    if clean:
        return clean[:255]
    fallback = (vendor or "Товар").strip() or "Товар"
    return fallback[:255]


def parse_product_tz_xlsx(content: bytes, *, filename: str) -> tuple[str, list[dict[str, Any]]]:
    lower = filename.lower()
    if not lower.endswith(".xlsx"):
        raise ProductTzImportError(
            "unsupported_file_type",
            "Поддерживаются только файлы Excel (.xlsx).",
        )
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ProductTzImportError(
            "unsupported_file_type", "Не удалось прочитать Excel."
        ) from exc

    try:
        sheet_name, header_row, cols = _find_tz_sheet(wb)
        ws = wb[sheet_name]
        expanded: dict[str, dict[int, str]] = {}
        for field in _EXPAND_FIELDS:
            col = cols.get(field)
            if col is not None:
                expanded[field] = _merged_values_by_row(ws, col, header_row)

        rows: list[dict[str, Any]] = []
        max_row = ws.max_row or header_row
        for r in range(header_row + 1, max_row + 1):
            vendor = expanded.get("vendor_article", {}).get(r)
            product_name = expanded.get("name", {}).get(r)
            sku_raw = expanded.get("sku", {}).get(r)
            wb_nm_id_raw = expanded.get("wb_nm_id", {}).get(r)
            size = expanded.get("size", {}).get(r)
            barcode_raw = expanded.get("barcode", {}).get(r)
            label_raw = expanded.get("label_barcode", {}).get(r)
            tz = expanded.get("tz", {}).get(r)
            wb_nm_id, wb_nm_id_error = _parse_wb_nm_id(wb_nm_id_raw)
            has_product_cells = any(
                (product_name, vendor, sku_raw, wb_nm_id_raw, size, barcode_raw, label_raw, tz)
            )
            if not has_product_cells:
                continue
            barcode = _resolve_barcode(barcode_raw=barcode_raw, label_raw=label_raw)
            rows.append(
                {
                    "row": r,
                    "name": product_name,
                    "vendor_article": vendor,
                    "sku": sku_raw,
                    "wb_nm_id": wb_nm_id,
                    "wb_nm_id_error": wb_nm_id_error,
                    "size": size,
                    "barcode": barcode,
                    "packaging_instructions": tz,
                    "declared_quantity": None,
                }
            )
        if not rows:
            raise ProductTzImportError(
                "empty_file", "Файл пустой — нет строк с артикулами."
            )
        return sheet_name, rows
    finally:
        wb.close()


async def _find_by_barcode_seller(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    seller_id: uuid.UUID,
    barcode: str,
) -> Product | None:
    stmt = select(Product).where(
        Product.tenant_id == tenant_id,
        Product.seller_id == seller_id,
        Product.wb_barcode == barcode,
    )
    res = await session.execute(stmt)
    return res.scalar_one_or_none()


async def _find_by_barcode_tenant(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    barcode: str,
) -> Product | None:
    stmt = select(Product).where(
        Product.tenant_id == tenant_id,
        Product.wb_barcode == barcode,
    )
    res = await session.execute(stmt)
    return res.scalar_one_or_none()


async def _find_by_sku(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    seller_id: uuid.UUID,
    sku_code: str,
) -> Product | None:
    # Артикул уникален внутри продавца, поэтому и «занят» он может быть только своим:
    # без фильтра по продавцу запрос вернул бы товар чужого юрлица и ложно объявил
    # артикул занятым (а после снятия тенантного ограничения — ещё и упал бы на
    # нескольких строках).
    stmt = select(Product).where(
        Product.tenant_id == tenant_id,
        Product.seller_id == seller_id,
        Product.sku_code == sku_code,
    )
    res = await session.execute(stmt)
    return res.scalars().first()


def _error_preview(
    *,
    row_no: int,
    wb_nm_id: int | None,
    vendor: str | None,
    size: str | None,
    barcode: str | None,
    name: str,
    sku: str,
    tz: str | None,
    code: str,
    msg: str,
    declared_quantity: int | None = 0,
) -> ProductTzRowPreview:
    return ProductTzRowPreview(
        row=row_no,
        wb_nm_id=wb_nm_id,
        vendor_article=vendor,
        size=size,
        barcode=barcode,
        name=name,
        sku_code=sku,
        packaging_instructions=tz,
        declared_quantity=declared_quantity,
        action="error",
        product_id=None,
        error_code=code,
        error_message=msg,
    )


async def build_product_tz_preview(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    *,
    seller_id: uuid.UUID,
    content: bytes,
    filename: str,
) -> ProductTzPreviewResult:
    seller = await session.get(Seller, seller_id)
    if seller is None or seller.tenant_id != tenant_id:
        raise ProductTzImportError("seller_not_found", "Селлер не найден.")

    sheet_name, raw_rows = parse_product_tz_xlsx(content, filename=filename)
    previews: list[ProductTzRowPreview] = []
    errors: list[ProductTzRowError] = []
    create_count = update_count = skip_count = error_count = 0
    seen_barcodes: set[str] = set()

    for raw in raw_rows:
        row_no = int(raw["row"])
        vendor = raw["vendor_article"] if isinstance(raw["vendor_article"], str) else None
        product_name = raw["name"] if isinstance(raw["name"], str) else None
        sku_raw = raw["sku"] if isinstance(raw["sku"], str) else None
        wb_nm_id = raw["wb_nm_id"] if isinstance(raw["wb_nm_id"], int) else None
        size = raw["size"] if isinstance(raw["size"], str) else None
        barcode = raw["barcode"] if isinstance(raw["barcode"], str) else None
        raw_tz = raw["packaging_instructions"]
        tz = raw_tz if isinstance(raw_tz, str) else None
        declared_quantity_raw = raw.get("declared_quantity")
        declared_quantity = (
            declared_quantity_raw if isinstance(declared_quantity_raw, int) else None
        )
        wb_nm_id_error_raw = raw.get("wb_nm_id_error")
        wb_nm_id_error = wb_nm_id_error_raw if isinstance(wb_nm_id_error_raw, str) else None
        name = _display_name(product_name, vendor)

        if wb_nm_id_error is not None:
            error_count += 1
            errors.append(
                ProductTzRowError(
                    row=row_no,
                    barcode=barcode,
                    code="invalid_wb_nm_id",
                    message=wb_nm_id_error,
                )
            )
            previews.append(
                _error_preview(
                    row_no=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor=vendor,
                    size=size,
                    barcode=barcode,
                    name=name,
                    sku="",
                    tz=tz,
                    code="invalid_wb_nm_id",
                    msg=wb_nm_id_error,
                    declared_quantity=declared_quantity,
                )
            )
            continue

        if not vendor:
            error_count += 1
            msg = "Нет артикула продавца (проверьте объединённые ячейки)."
            errors.append(
                ProductTzRowError(row=row_no, barcode=barcode, code="missing_vendor", message=msg)
            )
            previews.append(
                _error_preview(
                    row_no=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor=vendor,
                    size=size,
                    barcode=barcode,
                    name=name,
                    sku="",
                    tz=tz,
                    code="missing_vendor",
                    msg=msg,
                    declared_quantity=declared_quantity,
                )
            )
            continue

        if not barcode:
            error_count += 1
            msg = "Нет штрихкода (нужны цифры в «Штрихкод» или «Информация для этикетки»)."
            errors.append(
                ProductTzRowError(
                    row=row_no, barcode=None, code="missing_barcode", message=msg
                )
            )
            previews.append(
                _error_preview(
                    row_no=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor=vendor,
                    size=size,
                    barcode=None,
                    name=name,
                    sku="",
                    tz=tz,
                    code="missing_barcode",
                    msg=msg,
                    declared_quantity=declared_quantity,
                )
            )
            continue

        if barcode in seen_barcodes:
            error_count += 1
            msg = "Дубликат штрихкода в файле."
            errors.append(
                ProductTzRowError(
                    row=row_no,
                    barcode=barcode,
                    code="duplicate_barcode_in_file",
                    message=msg,
                )
            )
            previews.append(
                _error_preview(
                    row_no=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor=vendor,
                    size=size,
                    barcode=barcode,
                    name=name,
                    sku="",
                    tz=tz,
                    code="duplicate_barcode_in_file",
                    msg=msg,
                    declared_quantity=declared_quantity,
                )
            )
            continue
        seen_barcodes.add(barcode)

        sku = (sku_raw or "").strip()[:128] or _sku_for_row(
            vendor=vendor,
            size=size,
            barcode=barcode,
        )
        existing = await _find_by_barcode_seller(session, tenant_id, seller_id, barcode)
        if existing is not None:
            update_count += 1
            previews.append(
                ProductTzRowPreview(
                    row=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor_article=vendor,
                    size=size,
                    barcode=barcode,
                    name=name if product_name else existing.name,
                    sku_code=existing.sku_code,
                    packaging_instructions=tz,
                    declared_quantity=declared_quantity,
                    action="update",
                    product_id=existing.id,
                    error_code=None,
                    error_message=None,
                )
            )
            continue

        other = await _find_by_barcode_tenant(session, tenant_id, barcode)
        if other is not None:
            error_count += 1
            msg = "Штрихкод уже занят другим селлером в этом ФФ."
            errors.append(
                ProductTzRowError(
                    row=row_no,
                    barcode=barcode,
                    code="barcode_taken_other_seller",
                    message=msg,
                )
            )
            previews.append(
                _error_preview(
                    row_no=row_no,
                    wb_nm_id=wb_nm_id,
                    vendor=vendor,
                    size=size,
                    barcode=barcode,
                    name=name,
                    sku=sku,
                    tz=tz,
                    code="barcode_taken_other_seller",
                    msg=msg,
                    declared_quantity=declared_quantity,
                )
            )
            continue

        by_sku = await _find_by_sku(session, tenant_id, seller_id, sku)
        if by_sku is not None and (by_sku.wb_barcode or "").strip() != barcode:
            sku_alt = _sku_for_row(vendor=vendor, size=None, barcode=barcode)
            by_sku_alt = await _find_by_sku(session, tenant_id, seller_id, sku_alt)
            if by_sku_alt is not None and (by_sku_alt.wb_barcode or "").strip() != barcode:
                error_count += 1
                msg = "Артикул уже занят другим товаром."
                errors.append(
                    ProductTzRowError(
                        row=row_no, barcode=barcode, code="sku_taken", message=msg
                    )
                )
                previews.append(
                    _error_preview(
                        row_no=row_no,
                        wb_nm_id=wb_nm_id,
                        vendor=vendor,
                        size=size,
                        barcode=barcode,
                        name=name,
                        sku=sku,
                        tz=tz,
                        code="sku_taken",
                        msg=msg,
                        declared_quantity=declared_quantity,
                    )
                )
                continue
            sku = sku_alt

        create_count += 1
        previews.append(
            ProductTzRowPreview(
                row=row_no,
                wb_nm_id=wb_nm_id,
                vendor_article=vendor,
                size=size,
                barcode=barcode,
                name=name,
                sku_code=sku,
                packaging_instructions=tz,
                declared_quantity=declared_quantity,
                action="create",
                product_id=None,
                error_code=None,
                error_message=None,
            )
        )

    summary = ProductTzPreviewSummary(
        total=len(previews),
        create_count=create_count,
        update_count=update_count,
        skip_count=skip_count,
        error_count=error_count,
        declared_total=sum(
            row.declared_quantity or 0 for row in previews if row.action != "error"
        ),
    )
    return ProductTzPreviewResult(
        rows=tuple(previews),
        errors=tuple(errors),
        summary=summary,
        sheet_name=sheet_name,
    )


async def apply_product_tz_import(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    *,
    seller_id: uuid.UUID,
    content: bytes,
    filename: str,
    ignore_errors: bool = False,
) -> ProductTzApplyResult:
    preview = await build_product_tz_preview(
        session,
        tenant_id,
        seller_id=seller_id,
        content=content,
        filename=filename,
    )
    if preview.summary.error_count and not ignore_errors:
        raise ProductTzImportError(
            "row_errors",
            f"В файле есть ошибки строк ({preview.summary.error_count}).",
        )

    file_sha256 = hashlib.sha256(content).hexdigest()
    import_record = ProductTzImport(
        tenant_id=tenant_id,
        seller_id=seller_id,
        warehouse_id=None,
        warehouse_scope=_NO_WAREHOUSE_SCOPE,
        import_type=_IMPORT_TYPE,
        file_sha256=file_sha256,
        filename=filename[:512],
        declared_total=0,
        movement_count=0,
    )
    try:
        # First write starts the outer DB transaction before any per-row savepoints.
        # This is required for real all-or-nothing rollback on SQLite as well as Postgres.
        session.add(import_record)
        await session.flush()
    except IntegrityError as exc:
        await session.rollback()
        if not _is_idempotency_conflict(exc):
            raise
        return ProductTzApplyResult(
            created_count=0,
            updated_count=0,
            skipped_count=preview.summary.total,
            product_ids=(),
            summary=preview.summary,
            errors=preview.errors,
            added_quantity=0,
            movement_count=0,
            already_applied=True,
            warehouse_id=None,
        )

    created = 0
    updated = 0
    skipped = 0
    added_quantity = 0
    movement_count = 0
    product_ids: list[uuid.UUID] = []

    try:
        for row in preview.rows:
            if row.action == "error":
                skipped += 1
                continue
            if row.action == "update":
                assert row.product_id is not None
                async with session.begin_nested():
                    existing = await session.get(Product, row.product_id)
                    if existing is None:
                        skipped += 1
                        continue
                    existing.name = row.name
                    if row.size:
                        existing.wb_size = row.size
                    if row.wb_nm_id is not None:
                        existing.wb_nm_id = row.wb_nm_id
                    if row.vendor_article:
                        existing.wb_vendor_code = row.vendor_article
                    if row.packaging_instructions is not None:
                        await update_packaging_instructions(
                            session,
                            tenant_id,
                            existing.id,
                            packaging_instructions=row.packaging_instructions,
                            commit=False,
                        )
                    else:
                        await session.flush()
                    product_ids.append(existing.id)
                    updated += 1
                continue
            if row.action == "create":
                if not row.barcode:
                    skipped += 1
                    continue
                try:
                    async with session.begin_nested():
                        p = await create_product(
                            session,
                            tenant_id,
                            name=row.name,
                            sku_code=row.sku_code,
                            length_mm=None,
                            width_mm=None,
                            height_mm=None,
                            seller_id=seller_id,
                            wb_barcode=row.barcode,
                            wb_size=row.size,
                            wb_vendor_code=row.vendor_article,
                            packaging_instructions=row.packaging_instructions,
                            commit=False,
                        )
                        if row.wb_nm_id is not None:
                            p.wb_nm_id = row.wb_nm_id
                        product_ids.append(p.id)
                        created += 1
                except IntegrityError as exc:
                    err = str(getattr(exc, "orig", exc)).lower()
                    code = (
                        "barcode_taken"
                        if "wb_barcode" in err or "uq_products_tenant_wb_barcode" in err
                        else "sku_taken"
                    )
                    if ignore_errors:
                        skipped += 1
                        continue
                    raise ProductTzImportError(
                        code, "Конфликт уникальности при создании."
                    ) from exc
                except CatalogError as exc:
                    if ignore_errors and exc.code in {
                        "sku_taken",
                        "barcode_taken",
                    }:
                        skipped += 1
                        continue
                    raise ProductTzImportError(exc.code, str(exc)) from exc
        import_record.movement_count = movement_count
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    return ProductTzApplyResult(
        created_count=created,
        updated_count=updated,
        skipped_count=skipped,
        product_ids=tuple(product_ids),
        summary=preview.summary,
        errors=preview.errors,
        added_quantity=added_quantity,
        movement_count=movement_count,
        already_applied=False,
        warehouse_id=None,
    )
