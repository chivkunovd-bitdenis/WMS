#!/usr/bin/env python3
"""One-off owner transfer from Transfer.xlsx, sheet 24.08.

The default mode is read-only.  Applying requires both the approval token printed by
the latest dry-run and an explicit confirmation equal to the run id.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import sys
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.exc import IntegrityError

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT / "backend") not in sys.path:
    sys.path.insert(0, str(ROOT / "backend"))

from app.db.session import SessionLocal  # noqa: E402
from app.services.ownership_transfer_service import (  # noqa: E402
    OwnershipTransferError,
    OwnershipTransferInput,
    apply_ownership_transfer_plan,
    build_ownership_transfer_plan,
)

SHEET_NAME = "24.08"
KNOWN_SOURCE_SHA256 = "34491b265ad740e9a8a985a235836e591d613c455b9dad4886b6dabff80abe75"
EXPECTED_ACTIVE_ROWS = 54
EXPECTED_ACTIVE_QUANTITY = 2_794
EXCLUDED_J308_SKUS = frozenset({"J308-6", "J308-24", "J308-25"})
CORRECTED_TARGET_BARCODES = {
    ("F907-1", "40", "2041885503642"): "2041373819071",
}


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _quantity(value: object) -> int:
    text = _cell_text(value)
    if text is None:
        return 0
    try:
        parsed = int(text)
    except ValueError as exc:
        raise OwnershipTransferError(f"invalid_quantity:{text}") from exc
    return parsed


def _row_value(values: list[object], headers: dict[str | None, int], column: str) -> object:
    index = headers[column] - 1
    return values[index] if index < len(values) else None


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_transfer_workbook(
    path: Path,
    *,
    enforce_known_source: bool = True,
    enforce_expected_scope: bool = True,
) -> tuple[str, list[OwnershipTransferInput]]:
    checksum = source_sha256(path)
    if enforce_known_source and checksum != KNOWN_SOURCE_SHA256:
        raise OwnershipTransferError(
            f"unexpected_source_sha256:{checksum};expected:{KNOWN_SOURCE_SHA256}"
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise OwnershipTransferError("sheet_24.08_not_found")
    sheet = workbook[SHEET_NAME]
    headers = {
        _cell_text(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if _cell_text(cell.value)
    }
    required = {
        "Артикул продавца",
        "Размер",
        "Баркод Loviana",
        "Баркод Fashion",
        "Переносим шт.",
    }
    missing = sorted(required - headers.keys())
    if missing:
        raise OwnershipTransferError(f"missing_columns:{','.join(missing)}")

    inputs: list[OwnershipTransferInput] = []
    for row_number, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row_cells]
        sku = _cell_text(_row_value(values, headers, "Артикул продавца"))
        if sku is None:
            continue
        size = _cell_text(_row_value(values, headers, "Размер")) or ""
        source_barcode = _cell_text(_row_value(values, headers, "Баркод Loviana"))
        target_barcode = _cell_text(_row_value(values, headers, "Баркод Fashion"))
        quantity = _quantity(_row_value(values, headers, "Переносим шт."))
        excluded_reason = None
        if 2 <= row_number <= 24:
            excluded_reason = "rows_2_24_excluded"
        elif 55 <= row_number <= 59:
            excluded_reason = "rows_55_59_excluded"
        elif sku.upper() in EXCLUDED_J308_SKUS:
            excluded_reason = "j308_loafers_excluded_by_owner"
        elif target_barcode is None:
            excluded_reason = "target_barcode_missing"

        original_target_barcode = None
        corrected = CORRECTED_TARGET_BARCODES.get((sku.upper(), size, target_barcode or ""))
        if corrected is not None:
            original_target_barcode = target_barcode
            target_barcode = corrected
        inputs.append(
            OwnershipTransferInput(
                row_number=row_number,
                sku=sku,
                size=size,
                source_barcode=source_barcode,
                target_barcode=target_barcode,
                quantity=quantity,
                excluded_reason=excluded_reason,
                original_target_barcode=original_target_barcode,
            )
        )
    workbook.close()

    active = [item for item in inputs if item.excluded_reason is None]
    active_quantity = sum(item.quantity for item in active)
    if enforce_expected_scope and (
        len(active) != EXPECTED_ACTIVE_ROWS or active_quantity != EXPECTED_ACTIVE_QUANTITY
    ):
        raise OwnershipTransferError(
            "unexpected_active_scope:"
            f"rows={len(active)} expected={EXPECTED_ACTIVE_ROWS};"
            f"quantity={active_quantity} expected={EXPECTED_ACTIVE_QUANTITY}"
        )
    return checksum, inputs


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Dry-run/apply transfer Loviana → ООО Фэшн from Transfer.xlsx"
    )
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--tenant-id", required=True, type=uuid.UUID)
    parser.add_argument("--warehouse-id", required=True, type=uuid.UUID)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--source-seller", default="Loviana")
    parser.add_argument("--target-seller", default="ООО Фэшн")
    parser.add_argument("--report", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--approval-token")
    parser.add_argument("--confirm")
    return parser


def _emit_report(report: dict[str, object], path: Path | None) -> None:
    payload = json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True)
    print(payload)
    if path is not None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(payload + "\n", encoding="utf-8")


async def _run(args: argparse.Namespace) -> int:
    if not args.file.is_file():
        raise OwnershipTransferError(f"source_file_not_found:{args.file}")
    checksum, inputs = parse_transfer_workbook(args.file)
    if args.apply:
        if args.confirm != args.run_id:
            raise OwnershipTransferError("apply_requires_confirm_equal_to_run_id")
        if not args.approval_token:
            raise OwnershipTransferError("apply_requires_approval_token")

    async with SessionLocal() as session:
        try:
            plan = await build_ownership_transfer_plan(
                session,
                tenant_id=args.tenant_id,
                warehouse_id=args.warehouse_id,
                run_id=args.run_id,
                source_sha256=checksum,
                inputs=inputs,
                source_seller_name=args.source_seller,
                target_seller_name=args.target_seller,
                lock=args.apply,
            )
            report: dict[str, object] = {
                "mode": "apply" if args.apply else "dry-run",
                "source_file": str(args.file.resolve()),
                **plan.as_dict(),
            }
            if not args.apply:
                await session.rollback()
                _emit_report(report, args.report)
                return 1 if plan.blockers else 0

            request = await apply_ownership_transfer_plan(
                session,
                approved_token=args.approval_token,
                plan=plan,
            )
            await session.commit()
            report["applied"] = True
            report["inbound_request_id"] = str(request.id)
            report["inbound_document_number"] = request.document_number
            _emit_report(report, args.report)
            return 0
        except (IntegrityError, OwnershipTransferError):
            await session.rollback()
            raise


def main() -> int:
    args = _parser().parse_args()
    try:
        return asyncio.run(_run(args))
    except (IntegrityError, OwnershipTransferError) as exc:
        error: dict[str, Any] = {
            "applied": False,
            "error": str(exc),
            "mode": "apply" if args.apply else "dry-run",
        }
        print(json.dumps(error, ensure_ascii=False, indent=2), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
